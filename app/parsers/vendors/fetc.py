"""遠通報表（.xlsx）。

Layout：
- r1: 主欄名（37 欄）
- r2: 子欄名（補充 col 24-27 細項）
- r3+: 資料

主要欄位：
- col 9:  停車交易日期
- col 10: 繳款日期
- col 11: 繳款時間
- col 13: 實際繳費金額
- col 14: 繳款工具
- col 17: 繳費方式（信用卡）
- col 0:  停車交易序號
- col 31: 繳款訂單編號
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.parsers.base import BaseParser
from app.utils.vendor_dates import parse_datetime_loose


def _to_amount(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").strip())
        except (TypeError, ValueError):
            return None


def _normalize_payment(s: Any) -> str | None:
    if s is None:
        return None
    text = str(s).strip()
    return {
        "信用卡": "creditcard",
        "現金": "cash",
        "悠遊卡": "easycard",
    }.get(text, text or None)


class FetcParser(BaseParser):

    def parse(self, file_path: str, job_id: str, period: str) -> list[dict]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"找不到檔案：{file_path}")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 3:
            return []

        out: list[dict] = []
        for r in rows[2:]:
            if not r or all(v in (None, "") for v in r):
                continue
            amount = _to_amount(r[13] if len(r) > 13 else None)  # 實際繳費金額
            if amount is None or amount <= 0:
                continue
            tdate = parse_datetime_loose(r[9] if len(r) > 9 else None)  # 停車交易日期
            if tdate is None:
                tdate = parse_datetime_loose(r[10] if len(r) > 10 else None)
            if tdate is None:
                continue
            payment_type = _normalize_payment(r[17] if len(r) > 17 else None)
            transaction_id = (
                str(r[0]).strip() if r[0] not in (None, "") else None
            )
            raw = {
                "停車交易序號": transaction_id,
                "車號": str(r[4]) if len(r) > 4 and r[4] not in (None, "") else None,
                "進場時間": str(r[7]) if len(r) > 7 else None,
                "出場時間": str(r[8]) if len(r) > 8 else None,
                "停車交易日期": str(r[9]) if len(r) > 9 else None,
                "臨停交易金額合計": _to_amount(r[12]) if len(r) > 12 else None,
                "實際繳費金額": amount,
                "繳款工具": str(r[14]) if len(r) > 14 and r[14] not in (None, "") else None,
                "繳費方式": str(r[17]) if len(r) > 17 and r[17] not in (None, "") else None,
                "發票號碼": str(r[19]) if len(r) > 19 and r[19] not in (None, "") else None,
            }
            out.append({
                "job_id": job_id,
                "venue_code": None,
                "payment_type": payment_type,
                "transaction_date": tdate,
                "amount": amount,
                "transaction_id": (transaction_id or "")[:100] or None,
                "raw_data": json.dumps(raw, ensure_ascii=False, default=str),
            })
        return out
