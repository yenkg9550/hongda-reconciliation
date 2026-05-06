"""詮營報表（.xlsx）。

Layout：
- r1: 空
- r2: 標題列「APS自動繳費機收費記錄」
- r3+r4: 雙層 header（r4 是部分子欄如「時間」「狀態」「加值」）
- r5+: 資料

欄位（依 col index）：
- col 0:  序號
- col 1:  付費別
- col 2:  車號
- col 3:  設備名稱
- col 4:  進入日 (date)
- col 5:  進入時間 (time)
- col 6:  計價日 (date)
- col 7:  計價時間 (time)
- col 8:  票種
- col 9:  停車費
- col 10: 實收
- col 11: 發票號碼
- col 12: 發票狀態
- col 13: 自動加值
- col 14: 統一編號
- col 15: 紀錄時間
- col 16: 卡號

額外：偶有「手機條碼：xxx」備註行要 skip。
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.parsers.base import BaseParser
from app.utils.vendor_dates import parse_datetime_loose


def _to_amount(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").strip())
        except (TypeError, ValueError):
            return None


def _normalize_payment(s: Any) -> str | None:
    if s is None:
        return None
    text = str(s).strip()
    return {
        "現金": "cash",
        "悠遊卡": "easycard",
        "LinePay": "linepay",
        "LINE Pay": "linepay",
        "一卡通": "ipass",
    }.get(text, text or None)


class QuanyingParser(BaseParser):

    def parse(self, file_path: str, job_id: str, period: str) -> list[dict]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"找不到檔案：{file_path}")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 5:
            return []

        out: list[dict] = []
        for r in rows[4:]:  # 跳過 r1-r4 標題列
            if not r or all(v in (None, "") for v in r):
                continue
            # 略過備註行（如「手機條碼:xxx」col 12 有值但 col 1 空）
            if (len(r) <= 1 or r[1] in (None, "")) and len(r) > 12 and r[12] not in (None, ""):
                continue
            payment_raw = r[1] if len(r) > 1 else None
            if payment_raw in (None, ""):
                continue
            amount = _to_amount(r[10] if len(r) > 10 else None)  # 實收
            if amount is None or amount <= 0:
                continue
            tdate = parse_datetime_loose(r[6] if len(r) > 6 else None)  # 計價日
            if tdate is None:
                tdate = parse_datetime_loose(r[4] if len(r) > 4 else None)  # fallback 進入日
            if tdate is None:
                continue
            payment_type = _normalize_payment(payment_raw)
            transaction_id = (
                str(r[11]).strip()
                if len(r) > 11 and r[11] not in (None, "")
                else None
            )
            raw = {
                "序號": str(r[0]) if r[0] not in (None, "") else None,
                "付費別": str(payment_raw),
                "車號": str(r[2]) if len(r) > 2 and r[2] not in (None, "") else None,
                "設備名稱": str(r[3]) if len(r) > 3 and r[3] not in (None, "") else None,
                "進入日": str(r[4]) if len(r) > 4 else None,
                "進入時間": str(r[5]) if len(r) > 5 else None,
                "計價日": str(r[6]) if len(r) > 6 else None,
                "計價時間": str(r[7]) if len(r) > 7 else None,
                "票種": str(r[8]) if len(r) > 8 and r[8] not in (None, "") else None,
                "停車費": _to_amount(r[9]) if len(r) > 9 else None,
                "實收": amount,
                "發票號碼": transaction_id,
                "發票狀態": str(r[12]) if len(r) > 12 and r[12] not in (None, "") else None,
            }
            out.append({
                "job_id": job_id,
                "venue_code": None,
                "payment_type": payment_type,
                "transaction_date": tdate,
                "amount": amount,
                "transaction_id": (transaction_id or "")[:100] or None,
                "raw_data": json.dumps(raw, ensure_ascii=False, default=str),
            })
        return out
