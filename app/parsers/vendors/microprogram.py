"""微程式報表（.xlsx）。

兩種 schema 並存（用 sheet name 或 header 偵測）：
- 9 欄（sheet='計費查詢 - 明細'）：設備編號/車種/訊息名稱/卡號/靠卡時間/收費方式/付費卡號/金額/RRN
- 16 欄（sheet='Worksheet'）：設備編號/車種/車號/繳費時間/收費方式/付費卡號/實際支付/
        停車費/充電費/佔位費/點數折抵金額/總金額/RRN(CMAS檢索號)/卡片交易時間/卡片交易序號/備註
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.parsers.base import BaseParser
from app.utils.vendor_dates import parse_datetime_loose


_PAY_TYPE_MAP = {
    "現金": "cash",
    "悠遊卡": "easycard",
    "LinePay": "linepay",
    "LINE Pay": "linepay",
    "linepay": "linepay",
    "一卡通": "ipass",
    "一卡通Money": "ipass",
    "icash": "icash",
}


def _to_amount(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").strip())
        except (TypeError, ValueError):
            return None


def _normalize_payment(s: Any) -> str | None:
    if s is None:
        return None
    text = str(s).strip()
    return _PAY_TYPE_MAP.get(text, text or None)


class MicroprogramParser(BaseParser):

    def parse(self, file_path: str, job_id: str, period: str) -> list[dict]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"找不到檔案：{file_path}")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return []

        headers = list(rows[0])
        # 偵測 schema
        if "金額" in headers and "靠卡時間" in headers:
            return self._parse_short(rows, job_id, period)
        if "總金額" in headers and "繳費時間" in headers:
            return self._parse_long(rows, job_id, period)
        raise ValueError(f"微程式：無法識別 schema：headers={headers[:5]}")

    def _parse_short(self, rows: list, job_id: str, period: str) -> list[dict]:
        """9 欄版：col 4=靠卡時間, col 5=收費方式, col 7=金額。"""
        out: list[dict] = []
        for r in rows[1:]:
            if not r or all(v in (None, "") for v in r):
                continue
            amount = _to_amount(r[7] if len(r) > 7 else None)
            if amount is None or amount <= 0:
                continue
            tdate = parse_datetime_loose(r[4] if len(r) > 4 else None)
            if tdate is None:
                continue
            payment_type = _normalize_payment(r[5] if len(r) > 5 else None)
            transaction_id = (
                str(r[8]).strip() if len(r) > 8 and r[8] not in (None, "") else None
            )
            raw = {
                "設備編號": str(r[0]) if r[0] not in (None, "") else None,
                "車種": str(r[1]) if len(r) > 1 and r[1] not in (None, "") else None,
                "訊息名稱": str(r[2]) if len(r) > 2 and r[2] not in (None, "") else None,
                "卡號": str(r[3]) if len(r) > 3 and r[3] not in (None, "") else None,
                "靠卡時間": str(r[4]) if len(r) > 4 else None,
                "收費方式": str(r[5]) if len(r) > 5 and r[5] not in (None, "") else None,
                "付費卡號": str(r[6]) if len(r) > 6 and r[6] not in (None, "") else None,
                "金額": amount,
                "RRN": transaction_id,
            }
            out.append({
                "job_id": job_id,
                "venue_code": None,
                "payment_type": payment_type,
                "transaction_date": tdate,
                "amount": amount,
                "transaction_id": (transaction_id or "")[:100] or None,
                "raw_data": json.dumps(raw, ensure_ascii=False, default=str),
            })
        return out

    def _parse_long(self, rows: list, job_id: str, period: str) -> list[dict]:
        """16 欄版：col 3=繳費時間, col 4=收費方式, col 11=總金額。"""
        out: list[dict] = []
        for r in rows[1:]:
            if not r or all(v in (None, "") for v in r):
                continue
            amount = _to_amount(r[11] if len(r) > 11 else None)
            if amount is None or amount <= 0:
                continue
            tdate = parse_datetime_loose(r[3] if len(r) > 3 else None)
            if tdate is None:
                continue
            payment_type = _normalize_payment(r[4] if len(r) > 4 else None)
            transaction_id = (
                str(r[12]).strip() if len(r) > 12 and r[12] not in (None, "") else None
            )
            raw = {
                "設備編號": str(r[0]) if r[0] not in (None, "") else None,
                "車種": str(r[1]) if len(r) > 1 and r[1] not in (None, "") else None,
                "車號": str(r[2]) if len(r) > 2 and r[2] not in (None, "") else None,
                "繳費時間": str(r[3]) if len(r) > 3 else None,
                "收費方式": str(r[4]) if len(r) > 4 and r[4] not in (None, "") else None,
                "付費卡號": str(r[5]) if len(r) > 5 and r[5] not in (None, "") else None,
                "實際支付": _to_amount(r[6]) if len(r) > 6 else None,
                "總金額": amount,
                "RRN": transaction_id,
                "卡片交易時間": str(r[13]) if len(r) > 13 and r[13] not in (None, "") else None,
            }
            out.append({
                "job_id": job_id,
                "venue_code": None,
                "payment_type": payment_type,
                "transaction_date": tdate,
                "amount": amount,
                "transaction_id": (transaction_id or "")[:100] or None,
                "raw_data": json.dumps(raw, ensure_ascii=False, default=str),
            })
        return out
