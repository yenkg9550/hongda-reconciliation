"""永璽報表（.xls / .xlsx）。

注意：r2 列出 9 個欄名跟資料對不齊，**直接 hardcode 欄位 index**。

實際 17 欄資料 layout（依 r3+ 觀察）：
- col 0:  流水號
- col 1:  交易序號
- col 2:  入場時間
- col 3:  出場時間（也是繳費時間）
- col 4:  車號
- col 5:  費率名（如「費率1」）
- col 6:  空
- col 7:  時數
- col 8:  literal '時'
- col 9:  分數
- col 10: literal '分'
- col 11: 應收金額
- col 12: 折扣/折抵金額
- col 13: 實收金額
- col 14: 付款方式
- col 15: 統一編號
- col 16: 發票編號
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook

from app.parsers.base import BaseParser
from app.utils.vendor_dates import parse_datetime_loose


def _to_amount(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").strip())
        except (TypeError, ValueError):
            return None


def _normalize_payment(s: Any) -> str | None:
    if s is None:
        return None
    text = str(s).strip()
    return {
        "現金": "cash",
        "悠遊卡": "easycard",
        "LinePay": "linepay",
        "LINE Pay": "linepay",
        "一卡通": "ipass",
        "一卡通Money": "ipass",
        "信用卡": "creditcard",
    }.get(text, text or None)


def _read_rows(path: Path) -> list[list[Any]]:
    """讀 .xls 或 .xlsx 一致回傳 list of list。"""
    suffix = path.suffix.lower()
    if suffix == ".xls":
        wb = xlrd.open_workbook(str(path), ragged_rows=True)
        sheet = wb.sheet_by_index(0)
        return [sheet.row_values(r) for r in range(sheet.nrows)]
    elif suffix == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    raise ValueError(f"永璽：不支援副檔名 {suffix}")


class YongxiParser(BaseParser):

    def parse(self, file_path: str, job_id: str, period: str) -> list[dict]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"找不到檔案：{file_path}")

        rows = _read_rows(path)
        if len(rows) < 3:
            return []

        out: list[dict] = []
        for r in rows[2:]:  # r1 列印資訊、r2 欄名（不可靠），資料從 r3 起
            if not r or all(v in (None, "") for v in r):
                continue
            # 判斷資料行：col 0 是數字流水號 + col 13 是金額
            if len(r) < 14:
                continue
            try:
                serial = float(r[0])
            except (TypeError, ValueError):
                continue
            amount = _to_amount(r[13])  # 實收金額
            if amount is None or amount <= 0:
                continue
            tdate = parse_datetime_loose(r[3])  # 出場時間 = 繳費時間
            if tdate is None:
                tdate = parse_datetime_loose(r[2])  # fallback 入場時間
            if tdate is None:
                continue
            payment_type = _normalize_payment(r[14] if len(r) > 14 else None)
            transaction_id = (
                str(r[16]).strip()
                if len(r) > 16 and r[16] not in (None, "")
                else None
            )
            raw = {
                "流水號": int(serial),
                "交易序號": str(r[1]) if r[1] not in (None, "") else None,
                "入場時間": str(r[2]) if r[2] not in (None, "") else None,
                "出場時間": str(r[3]) if r[3] not in (None, "") else None,
                "車號": str(r[4]) if r[4] not in (None, "") else None,
                "費率": str(r[5]) if r[5] not in (None, "") else None,
                "應收金額": _to_amount(r[11]) if len(r) > 11 else None,
                "折抵金額": _to_amount(r[12]) if len(r) > 12 else None,
                "實收金額": amount,
                "付款方式": str(r[14]) if len(r) > 14 and r[14] not in (None, "") else None,
                "統一編號": str(r[15]) if len(r) > 15 and r[15] not in (None, "") else None,
                "發票編號": transaction_id,
            }
            out.append({
                "job_id": job_id,
                "venue_code": None,
                "payment_type": payment_type,
                "transaction_date": tdate,
                "amount": amount,
                "transaction_id": (transaction_id or "")[:100] or None,
                "raw_data": json.dumps(raw, ensure_ascii=False, default=str),
            })
        return out
