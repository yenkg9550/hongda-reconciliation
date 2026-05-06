"""阜爾車辯場報表（.xlsx）。

Layout：
- r1-r5: 表頭（含「營運收入明細表」、交易時間範圍）
- r6: 欄名（從 col 1 開始，col 0 是空）
- r7+: 資料（25 欄）

主要欄位（col index）：
- col 1:  項次
- col 2:  交易序號
- col 3:  車牌號碼
- col 5:  發票號碼
- col 10: 進場時間
- col 11: 繳費時間
- col 14: 出場時間
- col 15: 付款方式
- col 17: 應收金額
- col 18: 折扣金額
- col 19: 實收金額
- col 22: 交易狀態
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.parsers.base import BaseParser
from app.utils.vendor_dates import parse_datetime_loose


def _to_amount(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        try:
            return float(str(v).replace(",", "").strip())
        except (TypeError, ValueError):
            return None


def _normalize_payment(s: Any) -> str | None:
    if s is None:
        return None
    text = str(s).strip()
    return {
        "現金": "cash",
        "悠遊卡": "easycard",
        "LinePay": "linepay",
        "LINE Pay": "linepay",
        "一卡通": "ipass",
        "一卡通Money": "ipass",
        "信用卡": "creditcard",
    }.get(text, text or None)


class FuerCarParser(BaseParser):

    def parse(self, file_path: str, job_id: str, period: str) -> list[dict]:
        path = Path(file_path)
        if not path.is_file():
            raise FileNotFoundError(f"找不到檔案：{file_path}")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 7:
            return []

        out: list[dict] = []
        for r in rows[6:]:  # 跳過 r1-r6 標題
            if not r or all(v in (None, "") for v in r):
                continue
            if len(r) < 20:
                continue
            # 資料行：col 1 是項次（純數字字串）
            v1 = r[1]
            if v1 in (None, ""):
                continue
            try:
                int(str(v1).strip())
            except (TypeError, ValueError):
                continue
            amount = _to_amount(r[19])  # 實收金額
            if amount is None or amount <= 0:
                continue
            tdate = parse_datetime_loose(r[11])  # 繳費時間
            if tdate is None:
                tdate = parse_datetime_loose(r[14])  # fallback 出場時間
            if tdate is None:
                continue
            payment_type = _normalize_payment(r[15])
            transaction_id = (
                str(r[2]).strip() if len(r) > 2 and r[2] not in (None, "") else None
            )
            raw = {
                "項次": str(v1).strip(),
                "交易序號": transaction_id,
                "車牌號碼": str(r[3]).strip() if r[3] not in (None, "") else None,
                "發票號碼": str(r[5]).strip() if len(r) > 5 and r[5] not in (None, "") else None,
                "進場時間": str(r[10]) if len(r) > 10 else None,
                "繳費時間": str(r[11]) if len(r) > 11 else None,
                "出場時間": str(r[14]) if len(r) > 14 else None,
                "付款方式": str(r[15]).strip() if len(r) > 15 and r[15] not in (None, "") else None,
                "應收金額": _to_amount(r[17]) if len(r) > 17 else None,
                "折扣金額": _to_amount(r[18]) if len(r) > 18 else None,
                "實收金額": amount,
                "交易狀態": str(r[22]).strip() if len(r) > 22 and r[22] not in (None, "") else None,
            }
            out.append({
                "job_id": job_id,
                "venue_code": None,
                "payment_type": payment_type,
                "transaction_date": tdate,
                "amount": amount,
                "transaction_id": (transaction_id or "")[:100] or None,
                "raw_data": json.dumps(raw, ensure_ascii=False, default=str),
            })
        return out
