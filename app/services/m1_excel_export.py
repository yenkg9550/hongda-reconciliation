"""產 M1a Excel 報表（async/Mongo 版）。

格式跟 backend_20260506 的 bulk_pipeline.stage_export 一樣：
- 23 欄，含對帳狀態、銀行、reference 解析、vendor 對應、match 策略、bank_id、batch_id
- 凍結首列 + auto filter + 染色
- 加 Summary sheet
"""
from __future__ import annotations

import io
import logging
from collections import Counter
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.collections import M1_DETAILS, UPLOAD_JOBS, VENDOR_TX, VENUES

logger = logging.getLogger(__name__)


HEADERS = [
    "對帳狀態", "銀行交易日", "銀行金額", "銀行帳戶", "reference",
    "解析信心", "解析模式", "場代", "場名", "系統商", "支付方式",
    "vendor累計", "expected撥款", "差額", "差額%",
    "費率", "lag(D-N)", "累計N天", "match策略", "候選場代",
    "Vendor來源檔", "bank_id", "batch_id",
]

COL_WIDTHS = [11, 12, 12, 30, 38, 12, 22, 8, 14, 14, 14, 13, 13, 11, 9, 8, 8, 8, 18, 30, 50, 8, 18]

FILL_MATCHED = PatternFill("solid", fgColor="C6EFCE")
FILL_PARTIAL_GOOD = PatternFill("solid", fgColor="FFEB9C")
FILL_PARTIAL_FAIR = PatternFill("solid", fgColor="FFD8A8")
FILL_PARTIAL_POOR = PatternFill("solid", fgColor="FFCCCC")
FILL_UNMATCHED = PatternFill("solid", fgColor="E0E0E0")
FILL_HEADER = PatternFill("solid", fgColor="305496")
FONT_HEADER = Font(bold=True, color="FFFFFF")


async def _build_lookups(db: Any) -> tuple[dict, dict]:
    """venue_code → venue 文檔；venue_code → 已上傳 vendor 來源檔名集合。

    files_by_venue 走兩條路 fallback：
      (1) 先試 $lookup aggregation；
      (2) 失敗 / Atlas tier 限制時，改成手動 join（先抓 vendor_transactions 的 venue_code+job_id
          distinct、再批次查 upload_jobs.filename）。
    任何錯誤都會吞掉並 log，讓 export 本身不會 500。
    """
    venues: dict[str, dict] = {}
    try:
        async for v in db[VENUES].find({}):
            vc = v.get("venue_code") or v.get("_id")
            if vc:
                venues[vc] = v
    except Exception as exc:  # noqa: BLE001
        logger.exception("load venues failed: %s", exc)

    files_by_venue: dict[str, set[str]] = {}
    try:
        pipeline = [
            {"$match": {"venue_code": {"$ne": None}}},
            {"$group": {"_id": {"venue_code": "$venue_code", "job_id": "$job_id"}}},
            {"$lookup": {
                "from": UPLOAD_JOBS,
                "localField": "_id.job_id",
                "foreignField": "job_id",
                "as": "job",
            }},
            {"$unwind": "$job"},
            {"$project": {"_id": 0, "venue_code": "$_id.venue_code", "filename": "$job.filename"}},
        ]
        async for row in db[VENDOR_TX].aggregate(pipeline):
            vc = row.get("venue_code")
            if vc:
                files_by_venue.setdefault(vc, set()).add(row.get("filename") or "")
    except Exception as exc:  # noqa: BLE001
        logger.warning("vendor files $lookup failed (%s); fallback to manual join", exc)
        # Manual fallback：先抓 vendor 的 (venue_code, job_id) pairs，再去 upload_jobs 查 filename
        try:
            pairs: set[tuple[str, str]] = set()
            cursor = db[VENDOR_TX].find(
                {"venue_code": {"$ne": None}}, {"venue_code": 1, "job_id": 1, "_id": 0}
            )
            async for r in cursor:
                vc = r.get("venue_code"); jid = r.get("job_id")
                if vc and jid:
                    pairs.add((vc, jid))
            job_ids = list({jid for _, jid in pairs})
            jobname: dict[str, str] = {}
            if job_ids:
                async for j in db[UPLOAD_JOBS].find(
                    {"job_id": {"$in": job_ids}}, {"job_id": 1, "filename": 1, "_id": 0}
                ):
                    jobname[j["job_id"]] = j.get("filename") or ""
            for vc, jid in pairs:
                files_by_venue.setdefault(vc, set()).add(jobname.get(jid, ""))
        except Exception as exc2:  # noqa: BLE001
            logger.exception("manual file-by-venue lookup also failed: %s", exc2)

    return venues, files_by_venue


async def build_m1_workbook(
    db: Any, *, period_start: date, period_end: date
) -> Workbook:
    """從 m1_results 讀資料、產 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"M1a_{period_start.strftime('%Y-%m')}"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    venues, files_by_venue = await _build_lookups(db)

    cursor = db[M1_DETAILS].find(
        {
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
        }
    )
    rows = await cursor.to_list(length=200_000)

    # 排序：依狀態（matched/partial/unmatched）、再依差額
    status_order = {"matched": 0, "partial": 1, "unmatched": 2}
    rows.sort(key=lambda r: (
        status_order.get(r.get("status") or "unmatched", 9),
        float(r.get("diff") or 0),
    ))

    status_counts: Counter[str] = Counter()

    for r in rows:
        meta = r.get("meta") or {}
        status = r.get("status") or "unmatched"
        status_counts[status] += 1

        venue = venues.get(r.get("venue_code") or "") or {}
        bank_amount = _f(r.get("bank_amount") or r.get("actual_remit"))
        diff = _f(r.get("diff"))
        diff_pct = (abs(diff) / bank_amount * 100) if (diff is not None and bank_amount) else None

        cands = meta.get("classifier_candidates") or []
        vendor_files = ", ".join(sorted(files_by_venue.get(r.get("venue_code") or "", set())))

        ws.append([
            status,
            r.get("bank_transaction_date") or "",
            bank_amount,
            r.get("bank_filename") or "",
            r.get("reference") or "",
            meta.get("classifier_confidence", ""),
            meta.get("classifier_pattern", ""),
            r.get("venue_code") or "",
            venue.get("venue_name") or "",
            venue.get("vendor_code") or r.get("vendor_code") or "",
            r.get("payment_type") or "",
            _f(r.get("vendor_amount")),
            _f(r.get("expected_remit")),
            diff,
            round(diff_pct, 2) if diff_pct is not None else None,
            meta.get("rate"),
            meta.get("lag_days"),
            meta.get("accum_days"),
            meta.get("match_strategy", ""),
            ", ".join(cands) if cands else "",
            vendor_files,
            meta.get("bank_id") or r.get("bank_id") or "",
            r.get("reconcile_batch_id") or "",
        ])
        row_idx = ws.max_row
        abs_diff = abs(diff) if diff is not None else None
        if status == "matched":
            fill = FILL_MATCHED
        elif status == "partial":
            if abs_diff is not None and abs_diff < 10:
                fill = FILL_PARTIAL_GOOD
            elif abs_diff is not None and abs_diff < 100:
                fill = FILL_PARTIAL_FAIR
            else:
                fill = FILL_PARTIAL_POOR
        else:
            fill = FILL_UNMATCHED
        ws.cell(row=row_idx, column=1).fill = fill

        # 數字格式
        for col in (3, 12, 13, 14):
            ws.cell(row=row_idx, column=col).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=15).number_format = "0.00\\%"
        cell = ws.cell(row=row_idx, column=16)
        if cell.value is not None:
            cell.number_format = "0.00%"

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary", 0)
    ws2.append(["項目", "數量", "比例"])
    for cell in ws2[1]:
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
    total = sum(status_counts.values())
    for label, fill in [
        ("matched", FILL_MATCHED),
        ("partial", FILL_PARTIAL_GOOD),
        ("unmatched", FILL_UNMATCHED),
    ]:
        cnt = status_counts.get(label, 0)
        pct = f"{cnt / total * 100:.1f}%" if total else "0%"
        ws2.append([label, cnt, pct])
        ws2.cell(row=ws2.max_row, column=1).fill = fill
    ws2.append(["總計", total, "100%"])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    for i, w in enumerate([14, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    return wb


async def m1_workbook_bytes(
    db: Any, *, period_start: date, period_end: date
) -> bytes:
    wb = await build_m1_workbook(db, period_start=period_start, period_end=period_end)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _f(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None
