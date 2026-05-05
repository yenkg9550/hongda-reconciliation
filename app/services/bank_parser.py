"""銀行對帳單 parser。

先支援 2026資料/銀行 裡的主要銀行格式，輸出 MongoDB bank_entries documents。
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class BankParseError(Exception):
    """銀行檔案格式無法解析。"""


FULLWIDTH_TRANS = str.maketrans(
    "０１２３４５６７８９／－﹣；：　",
    "0123456789/--;: ",
)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\t", "").strip()


def _norm(value: Any) -> str:
    return _clean(value).translate(FULLWIDTH_TRANS).replace(" ", "")


def _amount(value: Any) -> Decimal | None:
    if value is None or _clean(value) == "":
        return None
    text = _clean(value).replace(",", "")
    try:
        amount = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return amount if amount > 0 else None


def _parse_date(value: Any) -> str | None:
    if value is None or _clean(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _norm(value)
    if re.fullmatch(r"\d{7}", text):
        year = int(text[:3]) + 1911
        return date(year, int(text[3:5]), int(text[5:7])).isoformat()

    match = re.search(r"(\d{4})/(\d{1,2})/(\d{1,2})", text)
    if match:
        y, m, d = map(int, match.groups())
        return date(y, m, d).isoformat()

    match = re.search(r"(\d{2,3})/(\d{1,2})/(\d{1,2})", text)
    if match:
        y, m, d = map(int, match.groups())
        if y < 1911:
            y += 1911
        return date(y, m, d).isoformat()
    return None


def _find_col(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    normalized = [_norm(h) for h in headers]
    for candidate in candidates:
        key = _norm(candidate)
        for idx, header in enumerate(normalized):
            if header == key:
                return idx
    return None


def _extract_venue_code(memo: str) -> str | None:
    text = _norm(memo)
    long_number = re.findall(r"\d{5,}", text)
    if long_number:
        return long_number[-1][-3:]
    matches = re.findall(r"(?<!\d)(\d{3})(?!\d)", text)
    if matches:
        return matches[-1]
    return None


def _infer_payment_source(source_name: str, description: str, memo: str) -> str | None:
    text = f"{source_name} {description} {memo}"
    if "阜爾" in text:
        return "fuer"
    if "LINE" in text or "跨行轉帳" in text or "跨行轉入" in text or "60558379" in text:
        return "linepay"
    if "悠遊" in text or "FXML入帳" in text or "國泰世華商業銀行受託信託財產專戶" in text:
        return "easycard"
    if "遠創" in text or "遠通" in text:
        return "fetc"
    if "現金" in text or "無摺" in text:
        return "cash"
    return None


def _rows_from_xlsx(file_bytes: bytes) -> list[list[Any]]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _rows_from_xls(file_bytes: bytes) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise BankParseError("此環境尚未安裝 xlrd，暫時無法解析 .xls 銀行檔") from exc

    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)
    rows: list[list[Any]] = []
    for r in range(sheet.nrows):
        out = []
        for c in range(sheet.ncols):
            value = sheet.cell_value(r, c)
            if sheet.cell_type(r, c) == xlrd.XL_CELL_DATE:
                value = datetime(*xlrd.xldate_as_tuple(value, book.datemode))
            out.append(value)
        rows.append(out)
    return rows


def _find_header_row(rows: list[list[Any]]) -> tuple[int, list[str]]:
    for idx, row in enumerate(rows[:12]):
        headers = [_clean(v) for v in row]
        normalized = {_norm(h) for h in headers}
        has_amount = bool(
            normalized & {"金額", "存入", "存入金額", "存款金額"}
        )
        has_date = bool(normalized & {"帳務日", "計息日", "交易日", "交易日期", "交易日期交易時間"})
        if has_amount and has_date:
            return idx, headers
    raise BankParseError("找不到銀行對帳單欄位列")


def parse_bank_entries(
    *,
    file_bytes: bytes,
    filename: str,
    source_name: str,
    job_id: str,
) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        rows = _rows_from_xlsx(file_bytes)
    elif suffix == ".xls":
        rows = _rows_from_xls(file_bytes)
    else:
        raise BankParseError("銀行上傳目前只支援 .xlsx / .xls")

    header_idx, headers = _find_header_row(rows)
    value_date_col = _find_col(headers, ("帳務日", "計息日", "交易日", "交易日期", "交易日期 交易時間"))
    amount_col = _find_col(headers, ("存入", "存入金額", "存款金額", "金額"))
    desc_col = _find_col(headers, ("摘要", "存摺摘要"))
    memo_col = _find_col(headers, ("備註/資金用途", "備註", "附言"))

    if value_date_col is None or amount_col is None:
        raise BankParseError("缺少必要欄位：入帳日期或存入金額")

    entries: list[dict[str, Any]] = []
    skipped = 0
    for row in rows[header_idx + 1 :]:
        if not any(_clean(v) for v in row):
            continue

        description = _clean(row[desc_col]) if desc_col is not None and desc_col < len(row) else None
        memo = _clean(row[memo_col]) if memo_col is not None and memo_col < len(row) else None

        if description == "續上一行" and entries:
            merged_memo = f"{entries[-1].get('memo_raw') or ''}{memo or ''}"
            entries[-1]["memo_raw"] = merged_memo
            entries[-1]["venue_code"] = _extract_venue_code(merged_memo)
            continue

        amount = _amount(row[amount_col] if amount_col < len(row) else None)
        value_date = _parse_date(row[value_date_col] if value_date_col < len(row) else None)
        if amount is None or value_date is None:
            skipped += 1
            continue

        payment_source = _infer_payment_source(source_name, description or "", memo or "")
        entries.append(
            {
                "job_id": job_id,
                "account_id": source_name,
                "value_date": value_date,
                "amount": str(amount),
                "description": description,
                "memo_raw": memo,
                "venue_code": _extract_venue_code(memo or "") if payment_source == "cash" else None,
                "payment_source": payment_source,
            }
        )

    if not entries:
        raise BankParseError(f"未解析到可用的銀行入帳資料（略過 {skipped} 列）")
    return entries
